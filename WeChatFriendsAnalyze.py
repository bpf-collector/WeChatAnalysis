'''
@Author       : bpf
@Date         : 2020-01-01 14:47:41
@LastEditTime : 2020-07-19 23:00:01
@Description  : This is a program which can analyze datas of WeChat friends.(微信好友分析)
@FilePath     : /Python/lib/try/WeChatAnalyze/WeChatAnalyze.py
'''
# -*- coding: utf-8 -*-

#  导入模块
import os
from wxpy import Bot
import openpyxl
from pyecharts import Map
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import jieba

class WeChatRobot:

    '''====================== 1. 获取微信好友信息 ======================'''
    def __init__(self, ToExcelFile="", ToCityFile="", ToMapProvinceFile="", ToMapCityFile=""):
        '''
        @description: 初始化机器人和其他参数
        @ToExcelFile        {str}   微信好友信息的表格文件路径(.xlsx)
        @ToCityFile         {str}   城市词云图的文件路径(.png/.jpg)
        @ToMapProvinceFile  {str}   省份地图的文件路径(.html)
        @ToMapCityFile      {str}   城市地图的文件路径(.html)
        '''
        # 初始化机器人，需要扫码
        self.bot = Bot()
        # 获取我所有的微信好友信息 - 存储基础信息(未处理)
        self.allFriends_Info = self.bot.friends()
        # 我的微信好友个数
        self.allFriends_Num = len(self.allFriends_Info)
        # 保存微信好友信息的表格文件路径(.xlsx)
        self.ExcelFile = ToExcelFile
        # 保存城市词云图的文件路径(.png/.jpg)
        self.WCOfCityFile = ToCityFile
        # 保存省份地图的文件路径(.html)
        self.MapProvinceFile = ToMapProvinceFile
        # 城市地图的文件路径(.html)
        self.MapCityFile = ToMapCityFile
        # 自动调用run方法，使得在实例化对象后自动运行其他函数
        self.run()

    '''====================== 2. 统计微信好友信息 ======================'''
    def getFriendsInfo(self):
        '''
        @description: 统计微信好友的全部信息
        '''
        # 存储微信好友的信息(经过信息处理的)
        self.friendsInfo = []
        # 定义列标题
        self.infoTitle = ['NickName', 'RemarkName', 'Sex', 'Province', 'City']
        for aFriend in self.allFriends_Info:
            # 获取昵称
            NickName = aFriend.raw.get(self.infoTitle[0], None)
            # 获取备注
            RemarkName = aFriend.raw.get(self.infoTitle[1], None)
            # 获取性别
            Sex = {1:"男", 2:"女", 0:"其他"}.get(aFriend.raw.get(self.infoTitle[2], None), None)
            # 获取省份
            Province = aFriend.raw.get(self.infoTitle[3], None)
            # 获取城市
            City = aFriend.raw.get(self.infoTitle[4], None)
            lisTmp = [NickName, RemarkName, Sex, Province, City]
            self.friendsInfo.append(lisTmp)
    
    '''====================== 3. 保存微信好友信息 ======================'''
    def saveFriendsInfoAsExcel(self, ExcelName):
        '''
        @description: 保存微信好友的信息到 Excel 表格中 
        '''
        # 生成openpyxl对象
        workbook = openpyxl.Workbook()
        # 激活表格
        sheet = workbook.active
        # 设置表格标题
        sheet.title = 'WeChatFriendsInfo'
        # 填充列标题到第一行
        for _ in range(len(self.infoTitle)):
            sheet.cell(row=1, column=_+1, value=self.infoTitle[_])
        # 填充微信好友信息，从第二行开始
        for i in range(self.allFriends_Num):
            for j in range(len(self.infoTitle)):
                sheet.cell(row=i+2, column=j+1, value=str(self.friendsInfo[i][j]))
        # 若文件名非空，则保存到该路径下
        if ExcelName != "":
            workbook.save(ExcelName)
            print(">>> Save WeChat friends' information successfully!")

    '''====================== 4. 分析微信好友信息 ======================'''
    def quiteAnalyzeFriendsInfo(self):
        ''' 分析数据，一步到位，直接了当 '''
        print(self.allFriends_Info.stats_text())
    
    '''====================== 5. 产生city词云图 ======================'''
    def creatWordCloudOfCity(self, CityName):
        '''
        @description: 使用获取的数据生成city词云图 
        '''
        # 获取所有的城市
        cityStr = ""
        for i in range(self.allFriends_Num):
            if self.friendsInfo[i][4] not in cityStr:
                cityStr += " " + self.friendsInfo[i][4]
        #jieba库精确模式分词
        wordlist = jieba.lcut(cityStr)
        cityStr = ' '.join(wordlist)
        # 加载背景图片
        #cloud_mask = np.array(Image.open(BackGroundFile))
        #设置词云图属性
        font = r'C:\Windows\Fonts\simfang.ttf' # 设置字体路径
        wc = WordCloud(
            background_color = 'black',     # 背景颜色
            #mask = cloud_mask,             # 背景图片
            max_words = 100,                # 设置最大显示的词云数
            font_path = font,               # 设置字体形式（在本机系统中）
            height = 300,                   # 图片高度
            width = 600,                    # 图片宽度
            max_font_size = 100,            # 字体最大值
            random_state = 100,             # 配色方案的种类
            )
        # 生成词云图
        myword = wc.generate(cityStr)
        #展示词云图
        plt.imshow(myword)
        plt.axis('off')
        plt.show()
        # 若文件名非空，则保存到该路径下
        if CityName != "":
            #保存词云图
            wc.to_file(CityName)
            print(">>> Creat WeChat wordcloud of city successfully!")
    
    '''===================== 6. 产生province地图 ====================='''
    def creatMapProvince(self, MapFile):
        '''
        @description: 使用获取的数据生成province地图
        '''
        # 获取所有省份
        provinceList, provinceNum = [], []
        for i in range(self.allFriends_Num):
            if self.friendsInfo[i][3] not in provinceList:
                provinceList.append(self.friendsInfo[i][3])
                provinceNum.append(0)
        for i in range(self.allFriends_Num):
            for j in range(len(provinceList)):
                if self.friendsInfo[i][3] == provinceList[j]:
                    provinceNum[j] += 1
        # 生成 Map
        map = Map("各省微信好友分布", width=1000, height=800)
        map.add("", provinceList, provinceNum, maptype="china", is_visualmap=True, visual_text_color='#000')
        # 若文件名非空，则保存到该路径下
        if MapFile != "":
            #map.show_config()
            map.render(MapFile)
            print(">>> Creat WeChat Map of Provinces seccessfully!")
    
    '''===================== 7. 产生city地图 ====================='''
    def creatMapCity(self, MapFile):
        '''
        @description: 使用获取的数据生成city地图
        '''
        # 获取所有省份
        CityList, CityNum = [], []
        for i in range(self.allFriends_Num):
            if self.friendsInfo[i][4] not in CityList:
                CityList.append(self.friendsInfo[i][4])
                CityNum.append(0)
        for i in range(self.allFriends_Num):
            for j in range(len(CityList)):
                if self.friendsInfo[i][4] == CityList[j]:
                    CityNum[j] += 1
        for i in range(len(CityList)):
            CityList[i] += '市'
        # 生成 Map
        map = Map("各市微信好友分布", width=1000, height=800)
        map.add("", CityList, CityNum, maptype="广东", is_visualmap=True, visual_text_color='#000')
        # 若文件名非空，则保存到该路径下
        if MapFile != "":
            map.render(MapFile)
            print(">>> Creat WeChat Map of Cities seccessfully!")
        
    '''===================== 8. 自动执行函数 ====================='''
    def run(self):
        # 获取微信好友信息
        self.getFriendsInfo()
        print(">>> Get WeChat friends' information successfully!")
        print(">>> Members:", self.allFriends_Num)
        # 保存微信好友信息
        self.saveFriendsInfoAsExcel(self.ExcelFile)
        # 分析微信好友信息
        self.quiteAnalyzeFriendsInfo()
        # 使用微信好友的 city 产生词云图
        self.creatWordCloudOfCity(self.WCOfCityFile)
        # 生成微信好友的 province 地图
        self.creatMapProvince(self.MapProvinceFile)
        # 生成微信好友的 city 地图
        self.creatMapCity(self.MapCityFile)

# 创建文件夹
def createDir(dirpath):
    # 判断创建文件夹
    if not os.path.exists(dirpath):
        os.makedirs(dirpath)


if __name__ == "__main__":
    dirpath = "./lib/try/WeChatAnalyze/mama"
    createDir(dirpath)
    ToExcelFile = dirpath+ "/FriendsInfo.xlsx"      # 微信好友信息的Excel表格保存路径
    ToPictureFile = dirpath+ "/CityWordCloud.png"   # 微信好友信息city词云图保存路径
    ToMapFileProvince = dirpath+ "/WeChatProvinceMap.html" # 微信好友信息province地图保存路径
    ToMapFileCity = dirpath+ "/WeChatCityMap.html"  # 微信好友信息city地图保存路径
    # WeChatRobot对象实例化
    robot = WeChatRobot(ToExcelFile, ToPictureFile, ToMapFileProvince, ToMapFileCity)